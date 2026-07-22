"""
exports.py — Exports fichiers (package thermie_debits).

Génère les livrables téléchargeables : base analysée (CSV), rapport de
qualité (CSV), synthèse (XLSX). Ces fonctions écrivent sur disque quand
output_dir est fourni ; l'app pourra aussi les appeler vers un buffer.
"""
from __future__ import annotations
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

from .core import _pnda


def rapport_structure(df):
    print("=" * 60)
    print("RAPPORT DE STRUCTURE")
    print("=" * 60)
    print(f"  Période    : {df['date'].min()} → {df['date'].max()}")
    print(f"  Jours      : {len(df)}")
    print(f"  NaN T_eau  : {df['T_eau_moy'].isna().sum()}")
    print(f"  NaN T_air  : {df['T_air'].isna().sum()}")
    print(f"  NaN Δ norm.: {df['Delta_TMm'].isna().sum()}")
    if "Q" in df.columns:
        print(f"  Débit Q    : {df['Q'].notna().sum()} jours ({df.attrs.get('base_debit','influencé')})")
    cols = ["T_eau_moy", "T_eau_max", "T_air", "Delta_TMm", "T_air_std", "Tmh"]
    if "Q" in df.columns: cols.append("Q")
    print(f"\n  Statistiques descriptives :")
    print(df[cols].describe().round(2))



def exporter_base(df, output_dir):
    cols = ["date", "T_eau_moy", "T_eau_max", "T_air", "Delta_TMm",
            "T_normale", "T_air_std", "Tmh"]
    labels = ["Date", "T_eau_moy_°C", "T_eau_max_°C", "T_air_°C",
              "Ecart_Normale", "T_normale_°C", "T_air_standardisee_°C", "Tmh_7j_°C"]
    for extra in ["Q", "Q_inf", "Q_desinf"]:
        if extra in df.columns:
            cols.append(extra); labels.append(extra + "_m3s")
    df[cols].rename(columns=dict(zip(cols, labels))).to_csv(
        f"{output_dir}base_analysee.csv", index=False, sep=";", decimal=",")
    print("✅ base_analysee.csv")



def exporter_rapport_qc(rapport, output_dir):
    if rapport is None or len(rapport) == 0:
        print("  (aucun enregistrement écarté par le QC)")
        return
    rapport.to_csv(f"{output_dir}rapport_qualite.csv",
                   index=False, sep=";", decimal=",")
    print(f"✅ rapport_qualite.csv ({len(rapport)} enreg. écartés)")


# ============================================================
# GRAPHIQUE 0 — Chronique thermique
# ============================================================

def exporter_synthese_xlsx(sens, vul, sgvt, contexte, nom, localisation, output_dir,
                            debit_res=None, cst_res=None, df_q_all=None, base="influencé"):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Synthèse thermique"
    def fill(h): return PatternFill("solid", fgColor=h.lstrip("#"))
    def bd(style="thin", color="CCCCCC"):
        s = Side(style=style, color=color); return Border(left=s, right=s, top=s, bottom=s)
    align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_l = Alignment(horizontal="left", vertical="center", wrap_text=True)
    font_small = Font(size=10)

    def hdr(row, col, text, bg, span=1, fg="FFFFFF"):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = fill(bg); c.font = Font(bold=True, color=fg); c.alignment = align_c
        c.border = bd("medium", "AAAAAA")
        if span > 1: ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
        return c
    def row_data(r, label, value, cat="", pts="", poids="", bg="FFFFFF"):
        for ci, v in enumerate([label, value, cat, pts, poids], 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill = fill(bg); c.alignment = align_l; c.border = bd(); c.font = font_small

    sg = sgvt; sr = sens; vr = vul
    r = 1
    ws.merge_cells(f"A{r}:E{r}")
    t = ws.cell(r, 1, f"Analyse Thermie & Débits — {nom}")
    t.font = Font(bold=True, size=13, color="FFFFFF"); t.fill = fill("2C3E50"); t.alignment = align_c
    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    s = ws.cell(r, 1, f"Sonde : {localisation}  |  Contexte : {contexte['label']}  |  "
                      f"Normales 1991–2020  |  Débit : {base}")
    s.fill = fill("D5D8DC"); s.alignment = align_c; s.font = font_small
    r += 2
    for ci, h in enumerate(["Paramètre", "Valeur", "Catégorie", "Pts", "Poids"], 1):
        hdr(r, ci, h, "2C3E50")
    r += 1

    hdr(r, 1, "SENSIBILITÉ", "2471A3", span=5); r += 1
    row_data(r, "Pente m", f"{sr['m']:.4f}", sr["sens_cat"], str(sg["pts_s"]),
             f"{int(round(sg.get('poids',{}).get('s',0.30)*100))}%", "EBF5FB"); r += 1
    row_data(r, "r Pearson / ρ Spearman", f"{sr['r_pearson']:.4f} / {sr['r_spearman']:.4f}", "—", "—", "—"); r += 1
    row_data(r, "R² — Variance expliquée", f"{sg['r2']:.4f} → {sg['r2_cat']}", "—", "—", "—", "D6EAF8"); r += 1
    row_data(r, "Robustesse |ρ−r|", f"{sg['robustesse']:.4f} → {sg['rob_cat']}", "—", "—", "—", "D5F5E3"); r += 1

    _pds = sg.get("poids", {"s": 0.30, "c": 0.40, "a": 0.30, "f": None})
    def _wp(w): return f"{int(round(w*100))}%" if w else "—"
    hdr(r, 1, f"VULNÉRABILITÉ ESTIVALE · {contexte['label']}", "1E8449", span=5); r += 1
    row_data(r, f"Tmh > {vr['seuil_chr']}°C (stress systémique)", f"{vr['pct_chr']:.1f}%",
             vr["cat_chr"], str(sg["pts_c"]), _wp(_pds["c"]), "EAFAF1"); r += 1
    row_data(r, f"Tmax > {vr['seuil_aigu']}°C (létalité systémique)", f"{vr['n_aigu']}j",
             vr["cat_aigu"], str(sg["pts_a"]), _wp(_pds["a"]), "EAFAF1"); r += 1

    # Composante fraie-croissance (affichée même si non évaluée, transparence)
    fr = sg.get("fraie")
    if fr is not None:
        hdr(r, 1, "FRAIE-CROISSANCE (hors étiage)", "B9770D", span=5); r += 1
        disponible = fr.get("disponible", False)
        esp_lim = fr.get("espece_limitante", "—")
        if disponible and sg.get("pts_f") is not None:
            pct_f = fr.get("pct_fraie", float("nan"))
            pct_txt = f"{pct_f:.1f}%" if pct_f == pct_f else "n/d"
            rec = fr.get("n_annees")
            val = pct_txt + (f"  ({rec} an)" if rec else "")
            row_data(r, f"Écart optimum · repère : {esp_lim}", val,
                     fr.get("cat_fraie", "—"), str(sg["pts_f"]), _wp(_pds["f"]),
                     "FEF9E7"); r += 1
        else:
            row_data(r, "Composante non évaluée (chronique lacunaire)", "—",
                     "SGVT sur 3 comp.", "—", "—", "FADBD8"); r += 1
        for s in fr.get("sous_indicateurs", []):
            if disponible and s["espece"] == esp_lim: continue
            fen = "-".join(map(str, s["fenetre"]))
            if s.get("evalue"):
                rec = f"  ({s['n_annees']} an)" if s.get("n_annees") else ""
                row_data(r, f"   ↳ {s['espece']} (fenêtre {fen})",
                         f"{s['pct']:.1f}%{rec}", s["cat"], f"P={s['P']}", "—",
                         "FCF3CF"); r += 1
            else:
                row_data(r, f"   ↳ {s['espece']} (fenêtre {fen})",
                         f"central {s['n_central']}j", "non évalué (mois central)",
                         "—", "—", "FADBD8"); r += 1

    ncomp = sg.get("composantes", 3)
    hdr(r, 1, f"SGVT — information d'appoint ({ncomp} composantes)", "6C3483", span=5); r += 1
    row_data(r, "SGVT", f"{sg['sgvt']:.2f} / 10", sg["interp"], "—", "—", "F4ECF7")
    for col in range(1, 6): ws.cell(r, col).font = Font(bold=True, size=11)
    r += 2

    if cst_res:
        hdr(r, 1, "DÉBITS DE RÉFÉRENCE THERMIQUE", "6C3483", span=5); r += 1
        # Q_thermie_bio — PRINCIPAL
        q_bio = cst_res.get("q_thermie_bio")
        if q_bio is not None:
            pnd = _pnda(df_q_all["Q"], q_bio) if df_q_all is not None else None
            row_data(r, "★ Q_thermie_bio (RÉSULTAT PRINCIPAL)", f"{q_bio:.3f} m³/s",
                     f"PNDA = {pnd:.0f}%" if pnd is not None else "—",
                     "×1,10", "—", "F5B7B1")
            for col in range(1, 6): ws.cell(r, col).font = Font(bold=True, size=11)
            r += 1
        else:
            row_data(r, "★ Q_thermie_bio", "Non applicable",
                     "Q*_vuln non détecté", "—", "—"); r += 1
        # Q_thermie_fonc — appoint
        q_fonc = cst_res.get("q_thermie_fonc")
        pnd_f = _pnda(df_q_all["Q"], q_fonc) if df_q_all is not None else None
        row_data(r, "Q_thermie_fonc (appoint)", f"{q_fonc:.3f} m³/s",
                 f"PNDA = {pnd_f:.0f}%" if pnd_f is not None else "—",
                 f"α={cst_res['alpha_fonc']*100:.0f}%", "—", "E8DAEF"); r += 1
        if debit_res:
            row_data(r, "  Q*_stat (AICc)", f"{debit_res['q_aicc']:.3f} m³/s",
                     f"ΔAICc={debit_res['delta_aicc_aicc']:.1f}", "—", "—"); r += 1
            row_data(r, "  Q*_vuln (base bio)",
                     f"{debit_res['q_vuln']:.3f} m³/s" if debit_res.get('q_vuln') else "non détecté",
                     f"stress={debit_res.get('q_vuln_chr')}", "—", "—"); r += 1

    for i, w in enumerate([40, 24, 30, 10, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    path = f"{output_dir}Synthese_Thermique_{nom.replace(' ', '_')}.xlsx"
    wb.save(path)
    print(f"✅ {Path(path).name}")
    return path


# ============================================================
# MAIN
# ============================================================
# ============================================================
# VOLET CLIMATIQUE (bonus) — contexte descriptif long terme
#   Réutilise les données déjà chargées (air+RR, eau, débit, normales).
#   Produit des figures de contextualisation : tendance thermique,
#   étiages, débit estival, température de l'eau, précipitations.
#   N'alimente PAS les débits de référence — purement contextuel.
